import openpyxl 
import util

def getdata():
	print("Enter the starting point :")
	s=input()
	print("Enter the End point: ")
	e=input()
	#print(s,e)
	return (s,e)

def printdata(state):
	totalcost=int(0)
	totaltime=int(0)
	seat_avail=float(0)
	#walk=int(0)
	direction=[]
	stations=[]
	for i in range (1,len(state)):
		totalcost=totalcost+state[i][6]
		totaltime=totaltime+state[i][4]
		seat_avail=seat_avail+state[i][5]
		#walk=walk+state[i][6]
		direction.append(state[i][3])
		stations.append(state[i][2])
	print("Total TIME: ",totaltime," min")
	print("Total COST: ",totalcost," Rs" )
	print("Seat Availability: ",seat_avail/(len(state)-1),"% chance")
	#print("Total walking distance: ",walk," meter")
	print("Metro stations to choose:")
	for i in range (0,len(direction)):
		print("->",stations[i],"(",direction[i],")",end=" ")
	print("\n")
	return



def wholedata(start,end):
	path ="Book1.xlsx"
	wb_obj = openpyxl.load_workbook(path)
	sheet_obj = wb_obj.active
	m_row = sheet_obj.max_row

	s=util.Stack()
	start_state=("","Nil",start,"",0,0,20,0);
	s.push([start_state])

	while not s.isEmpty():
		solution=s.pop()
		last_node=len(solution)-1
		current_state=(solution[last_node])
		current_pos=current_state[2]
		previous_pos=current_state[1]

		if(current_pos==end):
			printdata(solution)
			#choice(start,end)
			#for i in range (1,len(solution)):
			#	print (solution[i])
			#print ("\n")
		else:
			for i in range(1,m_row+1):
				previous = sheet_obj.cell(row = i, column = 1)
				starting = sheet_obj.cell(row = i, column = 2)
				ending = sheet_obj.cell(row = i, column = 3)
				color = sheet_obj.cell(row = i, column = 4)
				timing = sheet_obj.cell(row = i, column = 5)
				seat_avail = sheet_obj.cell(row = i, column = 6)
				walk = sheet_obj.cell(row = i, column = 7)
				cost = sheet_obj.cell(row = i, column = 8)
				start=current_pos
				prev=previous_pos
				if starting.value==start and previous.value==prev:
					path=solution[:]
					child_state=(previous.value, starting.value, ending.value, color.value, timing.value, seat_avail.value, cost.value)
					path.append(child_state)
					s.push(path)
					#print (starting.value, ending.value, color.value, timing.value, seat_avail.value, walk.value, cost.value)
	return


def costeffective(start,end):
	path ="Book1.xlsx"
	wb_obj = openpyxl.load_workbook(path)
	sheet_obj = wb_obj.active
	m_row = sheet_obj.max_row

	print("\n++ The Cost effective way ++\n")
	s=util.PriorityQueue()
	start_state=("","Nil",start,"",0,0,20,0);
	s.push([start_state],0)

	visited=[]
	count=int(0)

	while not s.isEmpty():
		solution=s.pop()
		last_node=len(solution)-1
		current_state=(solution[last_node])
		current_pos=current_state[2]
		previous_pos=current_state[1]

		if(current_pos==end):
			if solution not in visited:
				printdata(solution)
				count=count+1
				#print(count)
				#print(solution)
				visited.append(solution)
			#choice(start,end)
			#totalcost=int(0)
			#for i in range (1,len(solution)):
			#	print (solution[i])
			#	totalcost=totalcost+solution[i][7]
			#print ("\n Total cost is:",totalcost)
			# In the case, when you want whole table, remove return
			if count==3: 
				return
		else:
			for i in range(1,m_row+1):
				previous = sheet_obj.cell(row = i, column = 1)
				starting = sheet_obj.cell(row = i, column = 2)
				ending = sheet_obj.cell(row = i, column = 3)
				color = sheet_obj.cell(row = i, column = 4)
				timing = sheet_obj.cell(row = i, column = 5)
				seat_avail = sheet_obj.cell(row = i, column = 6)
				walk = sheet_obj.cell(row = i, column = 7)
				cost = sheet_obj.cell(row = i, column = 8)
				start=current_pos
				prev=previous_pos
				if starting.value==start and previous.value==prev:
					path=solution[:]
					child_state=(previous.value, starting.value, ending.value, color.value, timing.value, seat_avail.value, cost.value)
					path.append(child_state)

					totalcost=int(0)
					for i in path:
						totalcost=totalcost+i[6]
					s.push(path,totalcost)
	
	return		

def timeeffective(start,end):
	path ="Book1.xlsx"
	wb_obj = openpyxl.load_workbook(path)
	sheet_obj = wb_obj.active
	m_row = sheet_obj.max_row

	print("\n++ The Time effective way ++\n")
	s=util.PriorityQueue()
	start_state=("","Nil",start,"",0,0,20,0);
	s.push([start_state],0)

	count=int(0)
	visited=[]

	while not s.isEmpty():
		solution=s.pop()
		last_node=len(solution)-1
		current_state=(solution[last_node])
		current_pos=current_state[2]
		previous_pos=current_state[1]

		if(current_pos==end):
			if solution not in visited:
				printdata(solution)
				visited.append(solution)
				count=count+1
			#choice(start,end)
			#totaltime=int(0)
			#for i in range (1,len(solution)):
			#	print (solution[i])
			#	totaltime=totaltime+solution[i][4]
			#print ("\n Total time is :",totaltime)
			
			if count==3:
				return
		else:
			for i in range(1,m_row+1):
				previous = sheet_obj.cell(row = i, column = 1)
				starting = sheet_obj.cell(row = i, column = 2)
				ending = sheet_obj.cell(row = i, column = 3)
				color = sheet_obj.cell(row = i, column = 4)
				timing = sheet_obj.cell(row = i, column = 5)
				seat_avail = sheet_obj.cell(row = i, column = 6)
				walk = sheet_obj.cell(row = i, column = 7)
				cost = sheet_obj.cell(row = i, column = 8)
				start=current_pos
				prev=previous_pos
				if starting.value==start and previous.value==prev:
					path=solution[:]
					child_state=(previous.value, starting.value, ending.value, color.value, timing.value, seat_avail.value, cost.value)
					path.append(child_state)

					totalcost=int(0)
					for i in path:
						totalcost=totalcost+i[4]
					s.push(path,totalcost)
	return


def seatavailability(start,end):
	path ="Book1.xlsx"
	wb_obj = openpyxl.load_workbook(path)
	sheet_obj = wb_obj.active
	m_row = sheet_obj.max_row

	print("\n++ The Max seat availability way ++\n")
	s=util.PriorityQueue()
	start_state=("","Nil",start,"",0,0,20,0);
	s.push([start_state],0)

	count=int(0)
	visited=[]

	while not s.isEmpty():
		solution=s.pop()
		last_node=len(solution)-1
		current_state=(solution[last_node])
		current_pos=current_state[2]
		previous_pos=current_state[1]

		if(current_pos==end):
			if solution not in visited:
				printdata(solution)
				count=count+1
				visited.append(solution)
			#choice(start,end)
			#totalseat=float(0)
			#for i in range (1,len(solution)):
			#	print (solution[i])
			#	totalseat=totalseat+solution[i][5]
			#print ("\n Total cost is:",totalseat/(len(solution)-1),"%")
			# In the case, when you want whole table, remove return 
				
			if count==3: 
				return
		else:
			for i in range(1,m_row+1):
				previous = sheet_obj.cell(row = i, column = 1)
				starting = sheet_obj.cell(row = i, column = 2)
				ending = sheet_obj.cell(row = i, column = 3)
				color = sheet_obj.cell(row = i, column = 4)
				timing = sheet_obj.cell(row = i, column = 5)
				seat_avail = sheet_obj.cell(row = i, column = 6)
				walk = sheet_obj.cell(row = i, column = 7)
				cost = sheet_obj.cell(row = i, column = 8)
				start=current_pos
				prev=previous_pos
				if starting.value==start and previous.value==prev:
					path=solution[:]
					child_state=(previous.value, starting.value, ending.value, color.value, timing.value, seat_avail.value, cost.value)
					path.append(child_state)

					totalseat=float(0)  # Since it shows how much % seats are available, convert the whole in percentage
					for i in path:
						totalseat=totalseat+((100-i[5])/100) # Since, we give more priority to more seats and we are using min-Heap, so subtract from 100 
					s.push(path,totalseat)
	return


def choice(start,end):
	print("\nPress 1 to find way which is COST effective")
	print("Press 2 to find way which is TIME effective")
	print("Press 3 to find way which is SEAT effective")
	print("Press 4 for whole possible routes")
	print("Press 9 to change the data")
	print("Press any key to exit")
	t=input()
	if t=='1':
		costeffective(start,end)
	if t=='2':
		timeeffective(start,end)
	if t=='3':
		seatavailability(start,end)
	if t=='4':
		wholedata(start,end)
	if t=='9':
		(start,end)=getdata()
	if t=='1' or t=='2' or t=='3' or t=='4' or t=='9':
		choice(start,end)
	return


(start,end)=getdata()
choice(start,end)
#wholedata(start,end)
#costeffective(start,end)
#timeeffective(start,end)
#seatavailability(start,end)

